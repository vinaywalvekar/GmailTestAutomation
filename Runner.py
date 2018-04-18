from selenium import webdriver
import Gmail
import openpyxl
from openpyxl import load_workbook
import time

'''Handling XL sheet'''
wb = load_workbook('mysheet.xlsx')
sheet_name = wb['Sheet1']
rmax = sheet_name.max_row

'''Call Functions from Gmail file'''
driver = webdriver
g = Gmail.MyScript()

for row in range(2, 4):
    usn = sheet_name['A' + row.__str__() + '']
    pswd = sheet_name['B' + row.__str__() + '']
    g.Login(usn.value, pswd.value)
    time.sleep(5)
    if g.isSuccess == 'COMPOSE':
        loc = sheet_name['C' + row.__str__() + '']
        loc.value = 'Successful'
        time.sleep(2)
        g.Logout()
    else:
        loc2 = sheet_name['C' + row.__str__() + '']
        loc2.value = 'Unsuccessful Login'
        g.Logout()
    wb.save('mysheet.xlsx')


g.tearDown()
